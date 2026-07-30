import csv
import io
import json
import re
from datetime import datetime, timezone, timedelta

from fastapi import HTTPException, UploadFile, status
from openai import OpenAI
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models.exam import ExamQuestion
from app.models.question import Question
from app.models.question_version import QuestionVersion
from app.services.prompts import (
    PROMPT_VERSION,
    prompt_generate_equivalent,
    prompt_generate_topic,
    prompt_suggest_improvements,
    validate_ai_output,
)
from app.models.user import User
from app.schemas.question import QuestionCreate, QuestionUpdate


class QuestionService:
    @staticmethod
    def _build_filtered_query(
        db: Session, user: User,
        subject=None, topic=None, difficulty=None, question_type=None, search=None,
    ):
        q = db.query(Question).filter(Question.is_deleted == False)
        if user.role == "student":
            q = q.filter(Question.teacher_id == 0)
        elif user.role == "teacher":
            q = q.filter(Question.teacher_id == user.id)
        if subject:
            q = q.filter(Question.subject.ilike(f"%{subject}%"))
        if topic:
            q = q.filter(Question.topic.ilike(f"%{topic}%"))
        if difficulty:
            q = q.filter(Question.difficulty == difficulty)
        if question_type:
            q = q.filter(Question.question_type == question_type)
        if search:
            q = q.filter(
                Question.question_text.ilike(f"%{search}%")
                | Question.subject.ilike(f"%{search}%")
                | Question.topic.ilike(f"%{search}%")
            )
        return q.order_by(Question.created_at.desc())

    @staticmethod
    def create(db: Session, data: QuestionCreate, teacher: User) -> Question:
        options_json = json.dumps(data.options) if data.options else None
        question = Question(
            teacher_id=teacher.id,
            subject=data.subject,
            topic=data.topic,
            difficulty=data.difficulty,
            question_type=data.question_type,
            question_text=data.question_text,
            options=options_json,
            correct_answer=data.correct_answer,
            marks=data.marks,
            explanation=data.explanation,
            is_ai_generated=data.is_ai_generated,
            blooms_level=data.blooms_level,
            ai_model=data.ai_model,
            ai_prompt_used=data.ai_prompt_used,
            generation_source=data.generation_source,
        )
        db.add(question)
        db.commit()
        db.refresh(question)
        return question

    @staticmethod
    def get_all(
        db: Session, user: User,
        subject=None, topic=None, difficulty=None, question_type=None, search=None,
        page: int = 1, page_size: int = 20,
    ) -> dict:
        q = QuestionService._build_filtered_query(db, user, subject, topic, difficulty, question_type, search)
        total = q.count()
        items = q.offset((page - 1) * page_size).limit(page_size).all()
        return {
            "items": items,
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": max(1, (total + page_size - 1) // page_size),
        }

    @staticmethod
    def get_all_unpaginated(
        db: Session, user: User,
        subject=None, topic=None, difficulty=None, question_type=None, search=None,
    ):
        return QuestionService._build_filtered_query(db, user, subject, topic, difficulty, question_type, search).all()

    @staticmethod
    def get_by_id(db: Session, question_id: int, user: User) -> Question:
        question = db.query(Question).filter(Question.id == question_id, Question.is_deleted == False).first()
        if not question:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Question not found")
        if user.role != "admin" and question.teacher_id != user.id:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="You do not have access to this question")
        return question

    @staticmethod
    def _save_version(db: Session, question: Question, changed_by: int) -> None:
        max_ver = db.query(QuestionVersion).filter(QuestionVersion.question_id == question.id).order_by(QuestionVersion.version_number.desc()).first()
        next_ver = (max_ver.version_number + 1) if max_ver else 1
        snapshot = {
            "subject": question.subject,
            "topic": question.topic,
            "difficulty": question.difficulty,
            "question_type": question.question_type,
            "question_text": question.question_text,
            "options": question.options,
            "correct_answer": question.correct_answer,
            "marks": question.marks,
            "explanation": question.explanation,
        }
        version = QuestionVersion(
            question_id=question.id,
            version_number=next_ver,
            snapshot=json.dumps(snapshot),
            changed_by=changed_by,
        )
        db.add(version)

    @staticmethod
    def update(db: Session, question_id: int, data: QuestionUpdate, user: User) -> Question:
        question = QuestionService.get_by_id(db, question_id, user)
        if user.role != "admin" and question.teacher_id != user.id:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="You can only update your own questions")
        QuestionService._save_version(db, question, user.id)
        update_data = data.model_dump(exclude_unset=True)
        if "options" in update_data:
            update_data["options"] = json.dumps(update_data["options"]) if update_data["options"] else None
        for field, value in update_data.items():
            setattr(question, field, value)
        db.commit()
        db.refresh(question)
        return question

    @staticmethod
    def delete(db: Session, question_id: int, user: User) -> None:
        question = QuestionService.get_by_id(db, question_id, user)
        if user.role != "admin" and question.teacher_id != user.id:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="You can only delete your own questions")
        exam_links = db.query(ExamQuestion).filter(ExamQuestion.question_id == question_id).count()
        if exam_links > 0:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Cannot delete this question. It is linked to {exam_links} exam(s). Remove it from all exams first.",
            )
        question.is_deleted = True
        question.deleted_at = datetime.now(timezone.utc)
        db.commit()

    @staticmethod
    def duplicate(db: Session, question_id: int, user: User) -> Question:
        original = QuestionService.get_by_id(db, question_id, user)
        if user.role != "admin" and original.teacher_id != user.id:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="You can only duplicate your own questions")
        question = Question(
            teacher_id=original.teacher_id,
            subject=original.subject,
            topic=original.topic,
            difficulty=original.difficulty,
            question_type=original.question_type,
            question_text=original.question_text + " (copy)",
            options=original.options,
            correct_answer=original.correct_answer,
            marks=original.marks,
            explanation=original.explanation,
        )
        db.add(question)
        db.commit()
        db.refresh(question)
        return question

    @staticmethod
    def bulk_delete(db: Session, question_ids: list[int], user: User) -> dict:
        now = datetime.now(timezone.utc)
        questions = db.query(Question).filter(Question.id.in_(question_ids), Question.is_deleted == False).all()
        q_map = {q.id: q for q in questions}
        if user.role != "admin":
            q_map = {k: v for k, v in q_map.items() if v.teacher_id == user.id}
        exam_links = db.query(ExamQuestion.question_id, ExamQuestion.id).filter(ExamQuestion.question_id.in_(list(q_map.keys()))).all()
        linked_ids = {r[0] for r in exam_links}
        deleted = 0
        skipped = 0
        errors = []
        for qid in question_ids:
            question = q_map.get(qid)
            if not question:
                skipped += 1
                continue
            if qid in linked_ids:
                errors.append(f"Question {qid} is linked to {sum(1 for r in exam_links if r[0] == qid)} exam(s)")
                skipped += 1
                continue
            question.is_deleted = True
            question.deleted_at = now
            deleted += 1
        db.commit()
        return {"deleted": deleted, "skipped": skipped, "errors": errors}

    @staticmethod
    def create_bulk(db: Session, questions_data: list[QuestionCreate], teacher: User) -> list[Question]:
        created = []
        for data in questions_data:
            options_json = json.dumps(data.options) if data.options else None
            question = Question(
                teacher_id=teacher.id,
                subject=data.subject,
                topic=data.topic,
                difficulty=data.difficulty,
                question_type=data.question_type,
                question_text=data.question_text,
                options=options_json,
                correct_answer=data.correct_answer,
                marks=data.marks,
                explanation=data.explanation,
                is_ai_generated=data.is_ai_generated,
                blooms_level=data.blooms_level,
                ai_model=data.ai_model,
                ai_prompt_used=data.ai_prompt_used,
                generation_source=data.generation_source,
            )
            db.add(question)
            created.append(question)
        db.commit()
        for q in created:
            db.refresh(q)
        return created

    @staticmethod
    def export_csv(
        db: Session, user: User,
        subject=None, topic=None, difficulty=None, question_type=None, search=None,
    ) -> str:
        questions = QuestionService.get_all_unpaginated(db, user, subject, topic, difficulty, question_type, search)
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["subject", "topic", "difficulty", "question_type", "question_text", "options", "correct_answer", "marks", "explanation"])
        for q in questions:
            opts = ""
            if q.options:
                try:
                    parsed = json.loads(q.options)
                    if isinstance(parsed, list):
                        opts = "; ".join(parsed)
                except (json.JSONDecodeError, TypeError):
                    opts = q.options
            writer.writerow([
                q.subject, q.topic, q.difficulty, q.question_type,
                q.question_text, opts, q.correct_answer, q.marks,
                q.explanation or "",
            ])
        return output.getvalue()

    @staticmethod
    def export_json(
        db: Session, user: User,
        subject=None, topic=None, difficulty=None, question_type=None, search=None,
    ) -> str:
        questions = QuestionService.get_all_unpaginated(db, user, subject, topic, difficulty, question_type, search)
        result = []
        for q in questions:
            opts = None
            if q.options:
                try:
                    parsed = json.loads(q.options)
                    if isinstance(parsed, list):
                        opts = parsed
                except (json.JSONDecodeError, TypeError):
                    pass
            result.append({
                "id": q.id,
                "subject": q.subject,
                "topic": q.topic,
                "difficulty": q.difficulty,
                "question_type": q.question_type,
                "question_text": q.question_text,
                "options": opts,
                "correct_answer": q.correct_answer,
                "marks": q.marks,
                "explanation": q.explanation,
                "created_at": q.created_at.isoformat() if q.created_at else None,
            })
        return json.dumps(result, indent=2)

    @staticmethod
    def export_pdf(
        db: Session, user: User,
        subject=None, topic=None, difficulty=None, question_type=None, search=None,
    ) -> bytes:
        questions = QuestionService.get_all_unpaginated(db, user, subject, topic, difficulty, question_type, search)
        from fpdf import FPDF
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 16)
        pdf.cell(0, 12, "Question Bank Export", align="C")
        pdf.ln(18)
        pdf.set_font("Helvetica", "", 8)
        pdf.cell(0, 5, f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} | Total questions: {len(questions)}", align="C")
        pdf.ln(10)
        for i, q in enumerate(questions, 1):
            pdf.set_font("Helvetica", "B", 11)
            pdf.multi_cell(0, 6, f"Q{i}. {q.question_text}")
            pdf.ln(1)
            pdf.set_font("Helvetica", "", 9)
            meta = f"Subject: {q.subject} | Topic: {q.topic} | Difficulty: {q.difficulty} | Type: {q.question_type}"
            pdf.set_text_color(100, 100, 100)
            pdf.cell(0, 5, meta)
            pdf.set_text_color(0, 0, 0)
            pdf.ln(5)
            if q.question_type == "mcq" and q.options:
                try:
                    opts = json.loads(q.options)
                    if isinstance(opts, list):
                        for li, opt in enumerate(opts):
                            pdf.cell(0, 5, f"   {chr(65+li)}. {opt}")
                            pdf.ln(5)
                except (json.JSONDecodeError, TypeError):
                    pass
            pdf.set_font("Helvetica", "B", 9)
            pdf.cell(0, 5, f"Answer: {q.correct_answer}")
            pdf.ln(5)
            if q.explanation:
                pdf.set_font("Helvetica", "I", 9)
                pdf.multi_cell(0, 5, f"Explanation: {q.explanation}")
            pdf.set_font("Helvetica", "", 9)
            pdf.set_draw_color(200, 200, 200)
            pdf.line(10, pdf.get_y() + 2, 200, pdf.get_y() + 2)
            pdf.ln(6)
            if pdf.get_y() > 265:
                pdf.add_page()
        return bytes(pdf.output())

    @staticmethod
    def parse_excel_file(file: UploadFile, user: User) -> list[dict]:
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(
                status_code=status.HTTP_501_NOT_IMPLEMENTED,
                detail="Excel import requires openpyxl. Install it with: pip install openpyxl",
            )
        MAX_SIZE = 50 * 1024 * 1024
        ext = (file.filename or "").rsplit(".", 1)[-1].lower()
        if ext not in ("xlsx", "xls"):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only .xlsx or .xls files are supported")
        content = file.file.read(MAX_SIZE + 1)
        if len(content) > MAX_SIZE:
            raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail="File exceeds maximum size of 50 MB")
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel file must have a header row and at least one data row")
        headers_raw = [str(c).strip().lower() if c else "" for c in rows[0]]
        header_map = {
            "subject": ["subject", "sub"],
            "topic": ["topic"],
            "difficulty": ["difficulty", "diff"],
            "question_type": ["question_type", "type", "questiontype"],
            "question_text": ["question_text", "question", "text", "questiontext"],
            "options": ["options", "option"],
            "correct_answer": ["correct_answer", "answer", "correctanswer"],
            "marks": ["marks", "mark"],
            "explanation": ["explanation", "explain"],
        }
        def match_header(h):
            for field, aliases in header_map.items():
                if h in aliases:
                    return field
            return None
        headers = [match_header(h) for h in headers_raw]
        unknown_idx = next((i for i, h in enumerate(headers) if h is None), None)
        if unknown_idx is not None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Unknown column: '{rows[0][unknown_idx]}'. Valid headers: subject, topic, difficulty, question_type, question_text, options, correct_answer, marks, explanation.",
            )
        parsed = []
        errors = []
        for ri, row in enumerate(rows[1:], start=2):
            q = {}
            for ci, field in enumerate(headers):
                val = str(row[ci]).strip() if ci < len(row) and row[ci] is not None else ""
                if field == "options":
                    q[field] = [s.strip() for s in val.split(";") if s.strip()] if val else None
                elif field == "marks":
                    try:
                        q[field] = int(float(val)) if val else 1
                    except (ValueError, TypeError):
                        q[field] = 1
                else:
                    q[field] = val or None
            if not q.get("question_text"):
                continue
            if not q.get("subject"):
                errors.append(f"Row {ri}: subject is required")
            if not q.get("topic"):
                errors.append(f"Row {ri}: topic is required")
            if not q.get("correct_answer"):
                errors.append(f"Row {ri}: correct_answer is required")
            q.setdefault("difficulty", "medium")
            q.setdefault("question_type", "mcq")
            q.setdefault("marks", 1)
            parsed.append(q)
        wb.close()
        return {"parsed": parsed, "errors": errors}

    @staticmethod
    def parse_pdf_file(file: UploadFile, user: User) -> dict:
        try:
            import fitz
        except ImportError:
            raise HTTPException(
                status_code=status.HTTP_501_NOT_IMPLEMENTED,
                detail="PDF import requires PyMuPDF. Install it with: pip install pymupdf",
            )
        MAX_SIZE = 50 * 1024 * 1024
        ext = (file.filename or "").rsplit(".", 1)[-1].lower()
        if ext != "pdf":
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only .pdf files are supported")
        content = file.file.read(MAX_SIZE + 1)
        if len(content) > MAX_SIZE:
            raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail="File exceeds maximum size of 50 MB")
        doc = fitz.open(stream=content, filetype="pdf")
        text = ""
        for page in doc:
            text += page.get_text() + "\n"
        doc.close()
        import re
        blocks = re.split(r'\n\s*\n(?=\d+[\)\.])|\n(?=\d+[\)\.])', text.strip())
        blocks = [b.strip() for b in blocks if b.strip()]
        LABELS = {
            "subject": ["subject", "sub"],
            "topic": ["topic"],
            "difficulty": ["difficulty", "diff"],
            "question_type": ["question_type", "type", "questiontype"],
            "options": ["options", "option"],
            "correct_answer": ["correct_answer", "answer", "correctanswer", "ans"],
            "marks": ["marks", "mark"],
            "explanation": ["explanation", "explain"],
        }
        parsed = []
        errors = []
        for bi, block in enumerate(blocks):
            lines = block.split("\n")
            q = {"difficulty": "medium", "question_type": "mcq", "marks": 1}
            question_lines = []
            last_field = None
            for line in lines:
                s = line.strip()
                if not s:
                    continue
                matched = False
                for field, aliases in LABELS.items():
                    for alias in aliases:
                        prefix = alias + ":"
                        if s.lower().startswith(prefix):
                            val = s[len(prefix):].strip()
                            if field == "options":
                                q[field] = [x.strip() for x in val.split(";") if x.strip()] if val else None
                            elif field == "marks":
                                try:
                                    q[field] = int(float(val)) if val else 1
                                except (ValueError, TypeError):
                                    q[field] = 1
                            else:
                                q[field] = val or None
                            matched = True
                            break
                    if matched:
                        last_field = field
                        break
                if not matched:
                    if last_field == "options" and q.get("options"):
                        q["options"][-1] += " " + s
                        continue
                    opt = re.match(r'^[\(\)A-Da-d\s]+[\)\.]\s*(.+)', s)
                    if opt and q.get("question_type") == "mcq":
                        q.setdefault("options", [])
                        q["options"].append(opt.group(1).strip())
                    else:
                        question_lines.append(s)
            if question_lines:
                q["question_text"] = " ".join(question_lines)
            if q.get("question_text"):
                if not q.get("subject"):
                    errors.append(f"Block {bi+1}: subject is required")
                if not q.get("correct_answer"):
                    errors.append(f"Block {bi+1}: correct_answer is required")
                parsed.append(q)
            elif q.get("correct_answer") or q.get("difficulty") != "medium" or q.get("marks") != 1:
                if parsed:
                    prev = parsed[-1]
                    if q.get("correct_answer"):
                        prev["correct_answer"] = q["correct_answer"]
                    if q.get("difficulty") != "medium":
                        prev["difficulty"] = q["difficulty"]
                    if q.get("marks") != 1:
                        prev["marks"] = q["marks"]
                    if q.get("explanation"):
                        prev["explanation"] = q["explanation"]
        return {"parsed": parsed, "errors": errors}

    @staticmethod
    def _get_ai_client():
        if settings.GROQ_API_KEY:
            return OpenAI(base_url="https://api.groq.com/openai/v1", api_key=settings.GROQ_API_KEY)
        if settings.OPENAI_API_KEY:
            return OpenAI(api_key=settings.OPENAI_API_KEY)
        raise HTTPException(status_code=status.HTTP_501_NOT_IMPLEMENTED, detail="AI generation is not configured. Set GROQ_API_KEY or OPENAI_API_KEY in the server environment.")

    @staticmethod
    def _get_ai_model():
        return settings.GROQ_MODEL if settings.GROQ_API_KEY else settings.OPENAI_MODEL

    @staticmethod
    def generate_with_ai(data, db: Session = None) -> list[dict]:
        client = QuestionService._get_ai_client()
        model = QuestionService._get_ai_model()
        syllabus_context = ""
        if data.syllabus_ids and db:
            from app.models.syllabus import Syllabus
            entries = db.query(Syllabus).filter(Syllabus.id.in_(data.syllabus_ids), Syllabus.is_active == True).all()
            if entries:
                from app.services.prompts import _format_syllabus_context
                syllabus_context = _format_syllabus_context(entries)
        blooms_levels = list(data.blooms_levels) if data.blooms_levels else None
        prompt = prompt_generate_topic(
            subject=data.subject,
            topic=data.topic,
            difficulties=data.difficulties,
            types=data.question_types,
            count=data.count,
            marks=data.marks,
            blooms_levels=blooms_levels,
            syllabus_context=syllabus_context,
        )
        kwargs = {
            "model": model,
            "messages": [
                {"role": "system", "content": "You are an expert exam question writer. Return only valid JSON."},
                {"role": "user", "content": prompt},
            ],
            "temperature": 0.7,
        }
        use_openai = bool(settings.OPENAI_API_KEY) and not settings.GROQ_API_KEY
        if use_openai:
            kwargs["response_format"] = {"type": "json_object"}
        try:
            response = client.chat.completions.create(**kwargs)
        except Exception:
            raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="AI generation failed. Please try again later.")
        raw = response.choices[0].message.content
        try:
            parsed = json.loads(raw)
            questions = parsed if isinstance(parsed, list) else parsed.get("questions", [])
        except json.JSONDecodeError:
            raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="AI returned invalid JSON")
        if not questions:
            raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="AI returned no questions")
        validation_errors = validate_ai_output(questions)
        if validation_errors:
            raise HTTPException(
                status_code=status.HTTP_502_BAD_GATEWAY,
                detail="AI output validation failed: " + "; ".join(validation_errors),
            )
        for q in questions:
            q.setdefault("marks", data.marks)
            q.setdefault("explanation", None)
            q.setdefault("options", None)
            q.setdefault("blooms_level", None)
        return questions[:data.count]

    @staticmethod
    def suggest_improvements(db: Session, question_id: int, user: User) -> dict:
        question = QuestionService.get_by_id(db, question_id, user)
        client = QuestionService._get_ai_client()
        model = QuestionService._get_ai_model()
        prompt = f"""Review this exam question and suggest 3 specific improvements:

Question: {question.question_text}
Subject: {question.subject}
Topic: {question.topic}
Difficulty: {question.difficulty}
Type: {question.question_type}
Answer: {question.correct_answer}
Explanation: {question.explanation or 'None'}

Return ONLY valid JSON with a "suggestions" key containing an array of 3 strings, each being a specific, actionable improvement suggestion.
Valid JSON only, no markdown, no code fences."""
        kwargs = {
            "model": model,
            "messages": [
                {"role": "system", "content": "You are an expert educator reviewing exam questions. Return only valid JSON."},
                {"role": "user", "content": prompt},
            ],
            "temperature": 0.7,
        }
        use_openai = bool(settings.OPENAI_API_KEY) and not settings.GROQ_API_KEY
        if use_openai:
            kwargs["response_format"] = {"type": "json_object"}
        try:
            response = client.chat.completions.create(**kwargs)
        except Exception as e:
            raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="AI suggestion failed. Please try again later.")
        raw = response.choices[0].message.content
        try:
            parsed = json.loads(raw)
            suggestions = parsed if isinstance(parsed, list) else parsed.get("suggestions", [])
        except json.JSONDecodeError:
            raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="AI returned invalid JSON")
        return {"suggestions": suggestions[:5]}

    @staticmethod
    def bulk_update(db: Session, question_ids: list[int], update_data: dict, user: User) -> dict:
        questions = db.query(Question).filter(Question.id.in_(question_ids), Question.is_deleted == False).all()
        if user.role != "admin":
            questions = [q for q in questions if q.teacher_id == user.id]
        for question in questions:
            QuestionService._save_version(db, question, user.id)
            for field, value in update_data.items():
                if value is not None and field in ("subject", "topic", "difficulty", "question_type", "marks", "explanation"):
                    setattr(question, field, value)
        db.commit()
        return {"updated": len(questions), "skipped": len(question_ids) - len(questions), "errors": []}

    @staticmethod
    def bulk_duplicate(db: Session, question_ids: list[int], user: User) -> dict:
        originals = db.query(Question).filter(Question.id.in_(question_ids), Question.is_deleted == False).all()
        if user.role != "admin":
            originals = [q for q in originals if q.teacher_id == user.id]
        created_ids = []
        for original in originals:
            question = Question(
                teacher_id=original.teacher_id,
                subject=original.subject,
                topic=original.topic,
                difficulty=original.difficulty,
                question_type=original.question_type,
                question_text=original.question_text + " (copy)",
                options=original.options,
                correct_answer=original.correct_answer,
                marks=original.marks,
                explanation=original.explanation,
            )
            db.add(question)
            db.flush()
            created_ids.append(question.id)
        db.commit()
        return {"created": created_ids, "skipped": len(question_ids) - len(originals), "errors": []}

    @staticmethod
    def get_analytics(db: Session, user: User) -> dict:
        from sqlalchemy import case, func
        from sqlalchemy.sql import extract
        base = db.query(Question).filter(Question.is_deleted == False)
        if user.role == "teacher":
            base = base.filter(Question.teacher_id == user.id)
        total = base.count()
        diff_rows = base.with_entities(Question.difficulty, func.count(Question.id)).group_by(Question.difficulty).all()
        by_difficulty = {r[0]: r[1] for r in diff_rows}
        type_rows = base.with_entities(Question.question_type, func.count(Question.id)).group_by(Question.question_type).all()
        by_type = {r[0]: r[1] for r in type_rows}
        subj_rows = base.with_entities(Question.subject, func.count(Question.id)).group_by(Question.subject).all()
        by_subject = {r[0]: r[1] for r in subj_rows}
        thirty_days_ago = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
        recently_added = base.filter(Question.created_at >= thirty_days_ago).count()
        unused = base.outerjoin(ExamQuestion, ExamQuestion.question_id == Question.id).filter(ExamQuestion.id == None).count()
        if total == 0:
            avg_diff = "N/A"
        else:
            score_row = base.with_entities(func.sum(case((Question.difficulty == "hard", 3), (Question.difficulty == "easy", 1), else_=2))).first()
            total_score = score_row[0] or 0
            avg_score = round(total_score / total)
            avg_diff = {1: "easy", 2: "medium", 3: "hard"}.get(avg_score, "medium")
        total_ai_generated = base.filter(Question.is_ai_generated == True).count()
        ai_generated_saved = base.filter(Question.is_ai_generated == True, Question.generation_source == "generate").count()
        bloom_rows = base.with_entities(Question.blooms_level, func.count(Question.id)).filter(Question.blooms_level != None).group_by(Question.blooms_level).all()
        by_blooms_level = {r[0]: r[1] for r in bloom_rows} if bloom_rows else {}
        eight_weeks_ago = datetime.now(timezone.utc) - timedelta(days=56)
        ai_trend_base = base.filter(Question.is_ai_generated == True, Question.created_at >= eight_weeks_ago)
        trend_rows = ai_trend_base.with_entities(
            func.strftime("%Y-%W", Question.created_at).label("week"),
            func.count(Question.id),
        ).group_by("week").order_by("week").all()
        ai_generation_trend = [{"week": r[0], "count": r[1]} for r in trend_rows]
        recent_ai = base.filter(Question.is_ai_generated == True).order_by(Question.created_at.desc()).limit(5).all()
        recent_ai_activity = [{"id": q.id, "question_text": q.question_text[:100], "blooms_level": q.blooms_level, "created_at": q.created_at.isoformat() if q.created_at else None} for q in recent_ai]
        return {
            "total_questions": total,
            "by_difficulty": by_difficulty,
            "by_type": by_type,
            "by_subject": by_subject,
            "recently_added": recently_added,
            "unused": unused,
            "average_difficulty": avg_diff,
            "total_ai_generated": total_ai_generated,
            "ai_generated_saved": ai_generated_saved,
            "by_blooms_level": by_blooms_level,
            "ai_generation_trend": ai_generation_trend,
            "recent_ai_activity": recent_ai_activity,
        }

    @staticmethod
    def get_versions(db: Session, question_id: int, user: User) -> list[dict]:
        QuestionService.get_by_id(db, question_id, user)
        versions = db.query(QuestionVersion).filter(QuestionVersion.question_id == question_id).order_by(QuestionVersion.version_number.desc()).all()
        result = []
        for v in versions:
            try:
                snapshot = json.loads(v.snapshot)
            except (json.JSONDecodeError, TypeError):
                snapshot = {}
            result.append({
                "id": v.id,
                "question_id": v.question_id,
                "version_number": v.version_number,
                "snapshot": snapshot,
                "changed_by": v.changed_by,
                "created_at": v.created_at.isoformat() if v.created_at else None,
            })
        return result

    @staticmethod
    def generate_equivalent(db: Session, question_id: int, user: User, count: int = 1) -> list[dict]:
        question = QuestionService.get_by_id(db, question_id, user)
        client = QuestionService._get_ai_client()
        model = QuestionService._get_ai_model()

        options_str = ""
        if question.question_type == "mcq" and question.options:
            try:
                opts = json.loads(question.options)
                if isinstance(opts, list):
                    options_str = "\n".join(f"  {chr(65+i)}. {o}" for i, o in enumerate(opts))
            except (json.JSONDecodeError, TypeError):
                pass

        prompt = prompt_generate_equivalent(
            count=count,
            subject=question.subject,
            topic=question.topic,
            difficulty=question.difficulty,
            marks=question.marks,
            question_type=question.question_type,
            question_text=question.question_text,
            options_str=options_str,
            correct_answer=question.correct_answer,
            explanation=question.explanation,
            blooms_level=question.blooms_level,
        )

        kwargs = {
            "model": model,
            "messages": [
                {"role": "system", "content": "You are a teacher creating equivalent exam questions. Return only valid JSON."},
                {"role": "user", "content": prompt},
            ],
            "temperature": 0.8,
        }
        use_openai = bool(settings.OPENAI_API_KEY) and not settings.GROQ_API_KEY
        if use_openai:
            kwargs["response_format"] = {"type": "json_object"}

        try:
            response = client.chat.completions.create(**kwargs)
        except Exception:
            raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="AI generation failed. Please try again later.")

        raw = response.choices[0].message.content
        try:
            parsed = json.loads(raw)
            questions = parsed if isinstance(parsed, list) else parsed.get("questions", [])
        except json.JSONDecodeError:
            raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="AI returned invalid JSON")
        if not questions:
            raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="AI returned no questions")
        for q in questions:
            q.setdefault("marks", question.marks)
            q.setdefault("explanation", None)
            q.setdefault("options", None)
            q.setdefault("blooms_level", question.blooms_level)
        return questions[:count]

    @staticmethod
    def generate_equivalent_from_data(data: dict, count: int = 1) -> list[dict]:
        client = QuestionService._get_ai_client()
        model = QuestionService._get_ai_model()

        options_str = ""
        if data.get("question_type") == "mcq" and data.get("options"):
            opts = data["options"]
            if isinstance(opts, list):
                options_str = "\n".join(f"  {chr(65+i)}. {o}" for i, o in enumerate(opts))

        prompt = prompt_generate_equivalent(
            count=count,
            subject=data.get("subject", ""),
            topic=data.get("topic", ""),
            difficulty=data.get("difficulty", "medium"),
            marks=data.get("marks", 1),
            question_type=data.get("question_type", "mcq"),
            question_text=data.get("question_text", ""),
            options_str=options_str,
            correct_answer=data.get("correct_answer", ""),
            explanation=data.get("explanation"),
            blooms_level=data.get("blooms_level"),
        )

        kwargs = {
            "model": model,
            "messages": [
                {"role": "system", "content": "You are a teacher creating equivalent exam questions. Return only valid JSON."},
                {"role": "user", "content": prompt},
            ],
            "temperature": 0.8,
        }
        use_openai = bool(settings.OPENAI_API_KEY) and not settings.GROQ_API_KEY
        if use_openai:
            kwargs["response_format"] = {"type": "json_object"}

        try:
            response = client.chat.completions.create(**kwargs)
        except Exception:
            raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="AI generation failed. Please try again later.")

        raw = response.choices[0].message.content
        try:
            parsed = json.loads(raw)
            questions = parsed if isinstance(parsed, list) else parsed.get("questions", [])
        except json.JSONDecodeError:
            raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="AI returned invalid JSON")
        if not questions:
            raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="AI returned no questions")
        for q in questions:
            q.setdefault("marks", data.get("marks", 1))
            q.setdefault("explanation", None)
            q.setdefault("options", None)
            q.setdefault("blooms_level", data.get("blooms_level"))
        return questions[:count]

    @staticmethod
    def check_duplicates(db: Session, user: User, question_texts: list[str]) -> dict:
        from sqlalchemy import func
        results = []
        total_duplicates = 0
        for text in question_texts:
            normalized = text.strip().lower()[:200]
            existing = db.query(Question).filter(
                Question.teacher_id == user.id,
                Question.is_deleted == False,
                func.lower(func.substr(Question.question_text, 1, 200)) == normalized,
            ).first()
            if existing:
                results.append({
                    "text": text,
                    "is_duplicate": True,
                    "existing_question_id": existing.id,
                    "existing_question_text": existing.question_text,
                })
                total_duplicates += 1
            else:
                results.append({
                    "text": text,
                    "is_duplicate": False,
                    "existing_question_id": None,
                    "existing_question_text": None,
                })
        return {"results": results, "total_duplicates": total_duplicates}